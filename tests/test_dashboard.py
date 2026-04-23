"""Tests de dashboard basado en spec."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from exceland_factory.factory import build_product


class TestDashboardFromSpec:
    def test_caja_diaria_dashboard_has_kpis_from_spec(self, tmp_path: Path):
        """caja_diaria tiene kpis explícitos en el spec."""
        out = tmp_path / "caja_diaria.xlsx"
        result = build_product("caja_diaria", out)
        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        # Los KPIs deben estar en las filas 5, 7, 9, 11 (5 + i*2)
        # Columna B (índice 1) tiene las etiquetas
        assert ws["B5"].value == "Ingresos totales"
        assert ws["B7"].value == "Egresos totales"
        assert ws["B9"].value == "Flujo neto"
        assert ws["B11"].value == "Saldo acumulado"

    def test_precio_margen_dashboard_has_kpis_from_spec(self, tmp_path: Path):
        out = tmp_path / "precio_margen.xlsx"
        result = build_product("precio_margen", out)
        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        assert ws["B5"].value == "Precio de venta"
        assert ws["B7"].value == "Margen bruto %"
        assert ws["B9"].value == "Margen en pesos"
        assert ws["B11"].value == "Markup"

    def test_stock_control_dashboard_has_kpis_from_spec(self, tmp_path: Path):
        out = tmp_path / "stock_control.xlsx"
        result = build_product("stock_control", out)
        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        assert ws["B5"].value == "Alerta stock mínimo"
        assert ws["B7"].value == "Días de stock restante"
        assert ws["B9"].value == "Costo reposición promedio"
        assert ws["B11"].value == "Rotación inventario"

    def test_punto_equilibrio_dashboard_has_kpis_from_spec(self, tmp_path: Path):
        out = tmp_path / "punto_equilibrio.xlsx"
        result = build_product("punto_equilibrio", out)
        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        assert ws["B5"].value == "Punto equilibrio (uds)"
        assert ws["B7"].value == "Margen bruto %"
        assert ws["B9"].value == "Punto equilibrio ($)"
        assert ws["B11"].value == "Resultado neto"

    def test_dashboard_formulas_point_to_correct_refs(self, tmp_path: Path):
        """Las fórmulas del dashboard deben apuntar a las refs del spec."""
        out = tmp_path / "caja_diaria.xlsx"
        build_product("caja_diaria", out)

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        # Las fórmulas deben tener las referencias correctas
        # C5 = =MOTOR!C2
        formula_c5 = ws["C5"].value
        assert "MOTOR!C2" in str(formula_c5), f"Expected MOTOR!C2, got {formula_c5}"

    def test_fallback_legacy_dashboard_without_kpis(self, tmp_path: Path):
        """Un spec sin kpis debe usar el fallback legacy."""
        import textwrap
        from exceland_factory.workbook_builder import build_workbook
        from exceland_factory.validators import load_spec

        spec_file = tmp_path / "legacy_dashboard.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: legacy_test
                title: "Legacy Test"
                sheets:
                  - name: DATOS
                    type: input
                    fields:
                      - id: x
                        label: X
                        row: 5
                        col: 2
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    formulas: []
                  - name: DASHBOARD
                    type: dashboard
                    protected: true
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "legacy.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        # Fallback legacy debe tener los 4 KPIs originales
        assert ws["B5"].value == "Ingresos totales"
        assert ws["B7"].value == "Egresos totales"
        assert ws["B9"].value == "Resultado neto"
        assert ws["B11"].value == "Punto de equilibrio (uds)"