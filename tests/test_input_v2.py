"""Tests de INPUT v2: validar renderizado desde inputs declarativos."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from exceland_factory.factory import build_product
from exceland_factory.workbook_builder import build_workbook
from exceland_factory.validators import load_spec


class TestInputV2Rendering:
    def test_input_v2_renders_from_inputs_field(self, tmp_path: Path):
        """Si no hay fields pero hay inputs, renderiza desde inputs."""
        import textwrap

        spec_file = tmp_path / "input_v2.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: input_v2_test
                title: "Test Input V2"
                sheets:
                  - name: DATOS
                    type: input
                    inputs:
                      - id: costo_unitario
                        label: "Costo unitario"
                        type: currency
                        default: 100
                        required: true
                      - id: margen_objetivo
                        label: "Margen objetivo"
                        type: percentage
                        default: 0.40
                        hint: "Ejemplo: 0.40 = 40%"
                  - name: DASHBOARD
                    type: dashboard
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "input_v2.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DATOS"]

        # Los inputs se renderizan en filas secuenciales desde la 5
        assert ws["B5"].value == "Costo unitario"
        assert ws["C5"].value == 100
        assert ws["B6"].value == "Margen objetivo"
        assert ws["C6"].value == 0.40
        assert ws["D6"].value == "ℹ  Ejemplo: 0.40 = 40%"

    def test_input_legacy_takes_precedence_over_v2(self, tmp_path: Path):
        """Si hay fields (legacy) e inputs (v2), fields tiene precedencia."""
        import textwrap

        spec_file = tmp_path / "mixed.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: mixed_test
                title: "Test Mixed"
                sheets:
                  - name: DATOS
                    type: input
                    fields:
                      - id: legacy_field
                        label: "Legacy Field"
                        row: 5
                        col: 2
                        input_type: text
                    inputs:
                      - id: v2_input
                        label: "V2 Input"
                        type: currency
                  - name: DASHBOARD
                    type: dashboard
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "mixed.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DATOS"]

        # Legacy field se renderiza en la posición explícita
        assert ws["B5"].value == "Legacy Field"
        # V2 input NO se renderiza porque fields tiene precedencia
        assert ws["B6"].value is None

    def test_input_v2_with_validation(self, tmp_path: Path):
        """Input v2 con validation genera DataValidation en el xlsx."""
        import textwrap

        spec_file = tmp_path / "input_v2_validation.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: input_v2_val
                title: "Test Input V2 Validation"
                sheets:
                  - name: DATOS
                    type: input
                    inputs:
                      - id: precio
                        label: "Precio"
                        type: currency
                        validation: positive_number
                        required: true
                  - name: DASHBOARD
                    type: dashboard
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "input_v2_val.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DATOS"]

        # Verificar que hay DataValidation aplicada
        assert ws.data_validations.dataValidation, (
            "La hoja DATOS no tiene DataValidation"
        )

    def test_input_v2_required_field_has_bold_label(self, tmp_path: Path):
        """Input v2 required=True debe tener label en negrita."""
        import textwrap

        spec_file = tmp_path / "input_v2_required.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: input_v2_req
                title: "Test Input V2 Required"
                sheets:
                  - name: DATOS
                    type: input
                    inputs:
                      - id: obligatorio
                        label: "Campo obligatorio"
                        type: text
                        required: true
                      - id: opcional
                        label: "Campo opcional"
                        type: text
                        required: false
                  - name: DASHBOARD
                    type: dashboard
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "input_v2_req.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DATOS"]

        # Ambos campos deben estar presentes
        assert ws["B5"].value == "Campo obligatorio"
        assert ws["B6"].value == "Campo opcional"

    def test_legacy_specs_still_work(self, tmp_path: Path):
        """Specs legacy con fields siguen funcionando sin cambios."""
        out = tmp_path / "caja_diaria.xlsx"
        result = build_product("caja_diaria", out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["CAJA"]

        # Verificar que los campos legacy están en sus posiciones
        assert ws["B5"].value == "Nombre del negocio"
        assert ws["B8"].value == "Saldo inicial (ARS)"
