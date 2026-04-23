"""Tests de postprocess: verifica que DataValidation se aplica al xlsx final."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from exceland_factory.factory import build_product


class TestDataValidationApplied:
    def test_caja_diaria_has_validations(self, tmp_path: Path):
        """caja_diaria tiene campos con validation → el xlsx debe tener DataValidation."""
        out = tmp_path / "caja_diaria.xlsx"
        result = build_product("caja_diaria", out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["CAJA"]
        assert ws.data_validations.dataValidation, (
            "La hoja CAJA no tiene DataValidation — postprocess no se ejecutó"
        )

    def test_precio_margen_has_validations(self, tmp_path: Path):
        """precio_margen tiene campos con validation → el xlsx debe tener DataValidation."""
        out = tmp_path / "precio_margen.xlsx"
        result = build_product("precio_margen", out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DATOS"]
        assert ws.data_validations.dataValidation, (
            "La hoja DATOS no tiene DataValidation"
        )

    def test_punto_equilibrio_has_validations(self, tmp_path: Path):
        out = tmp_path / "punto_equilibrio.xlsx"
        result = build_product("punto_equilibrio", out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DATOS"]
        assert ws.data_validations.dataValidation, (
            "La hoja DATOS no tiene DataValidation"
        )

    def test_validation_cell_refs_are_correct(self, tmp_path: Path):
        """Verifica que las referencias de celda de las validaciones apuntan a celdas reales."""
        out = tmp_path / "precio_margen.xlsx"
        build_product("precio_margen", out)

        wb = load_workbook(str(out))
        ws = wb["DATOS"]

        # Recolectar todas las sqref de las validaciones
        validated_cells = set()
        for dv in ws.data_validations.dataValidation:
            validated_cells.add(str(dv.sqref))

        # costo_unitario está en row=7, col=2 → columna C → C7
        assert any("C7" in ref for ref in validated_cells), (
            f"No se encontró validación en C7 (costo_unitario). Refs: {validated_cells}"
        )

    def test_validation_type_is_correct(self, tmp_path: Path):
        """La validación positive_number debe ser de tipo decimal con operator greaterThan."""
        out = tmp_path / "precio_margen.xlsx"
        build_product("precio_margen", out)

        wb = load_workbook(str(out))
        ws = wb["DATOS"]

        # Buscar la validación de C7 (costo_unitario → positive_number)
        dv_for_c7 = None
        for dv in ws.data_validations.dataValidation:
            if "C7" in str(dv.sqref):
                dv_for_c7 = dv
                break

        assert dv_for_c7 is not None, "No se encontró validación para C7"
        assert dv_for_c7.type == "decimal", f"Tipo esperado 'decimal', got '{dv_for_c7.type}'"
        assert dv_for_c7.operator == "greaterThan", (
            f"Operador esperado 'greaterThan', got '{dv_for_c7.operator}'"
        )

    def test_build_still_succeeds_after_postprocess(self, tmp_path: Path):
        """El BuildResult debe ser success=True incluso con postprocess activo."""
        out = tmp_path / "stock_control.xlsx"
        result = build_product("stock_control", out)

        assert result.success, f"Build falló tras postprocess: {result.error}"
        assert out.exists()

    def test_spec_without_validations_skips_postprocess(self, tmp_path: Path):
        """Un spec sin campos con validation no debe fallar ni agregar validaciones."""
        import textwrap
        from exceland_factory.workbook_builder import build_workbook
        from exceland_factory.validators import load_spec

        spec_file = tmp_path / "no_validations.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: no_val_test
                title: "Sin validaciones"
                sheets:
                  - name: DATOS
                    type: input
                    fields:
                      - id: nombre
                        label: Nombre
                        row: 5
                        col: 2
                        input_type: text
                  - name: DASHBOARD
                    type: dashboard
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "no_val.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"
        assert out.exists()

        wb = load_workbook(str(out))
        ws = wb["DATOS"]
        # Sin validaciones en el spec → no debe haber DataValidation
        assert not ws.data_validations.dataValidation
