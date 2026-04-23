"""Tests de ENGINE v2: validar renderizado desde derived declarativos."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from exceland_factory.workbook_builder import build_workbook
from exceland_factory.validators import load_spec


class TestEngineV2Rendering:
    def test_engine_v2_renders_from_derived_field(self, tmp_path: Path):
        """Si no hay formulas pero hay derived, renderiza desde derived."""
        import textwrap

        spec_file = tmp_path / "engine_v2.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: engine_v2_test
                title: "Test Engine V2"
                sheets:
                  - name: DATOS
                    type: input
                    inputs:
                      - id: costo_unitario
                        label: "Costo unitario"
                        type: currency
                        default: 100
                      - id: margen_objetivo
                        label: "Margen objetivo"
                        type: percentage
                        default: 0.40
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    derived:
                      - id: precio_venta
                        formula: precio_venta_con_margen
                        inputs:
                          costo_unitario: costo_unitario
                          margen_objetivo: margen_objetivo
                      - id: margen_bruto
                        formula: margen_bruto
                        inputs:
                          precio_venta: precio_venta
                          costo_unitario: costo_unitario
                  - name: DASHBOARD
                    type: dashboard
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "engine_v2.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["MOTOR"]

        # Verificar que las fórmulas se escribieron
        assert ws["B1"].value == "ID"
        assert ws["C1"].value == "FÓRMULA"
        
        # Primera fórmula: precio_venta
        assert ws["B2"].value == "precio_venta"
        # La fórmula debe referenciar DATOS!C5 (costo_unitario) y DATOS!C6 (margen_objetivo)
        formula_pv = str(ws["C2"].value)
        assert "DATOS!C5" in formula_pv or "DATOS!C6" in formula_pv

        # Segunda fórmula: margen_bruto
        assert ws["B3"].value == "margen_bruto"
        # La fórmula debe referenciar MOTOR!C2 (precio_venta) y DATOS!C5 (costo_unitario)
        formula_mb = str(ws["C3"].value)
        assert "MOTOR!C2" in formula_mb

    def test_engine_legacy_takes_precedence_over_v2(self, tmp_path: Path):
        """Si hay formulas (legacy) y derived (v2), formulas tiene precedencia."""
        import textwrap

        spec_file = tmp_path / "mixed_engine.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: mixed_engine_test
                title: "Test Mixed Engine"
                sheets:
                  - name: DATOS
                    type: input
                    fields:
                      - id: x
                        label: "X"
                        row: 5
                        col: 2
                        input_type: text
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    formulas:
                      - id: legacy_formula
                        formula_ref: margen_bruto
                        bindings:
                          precio_venta: DATOS!C5
                          costo_unitario: DATOS!C6
                    derived:
                      - id: v2_formula
                        formula: margen_bruto
                        inputs:
                          precio_venta: x
                          costo_unitario: x
                  - name: DASHBOARD
                    type: dashboard
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "mixed_engine.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["MOTOR"]

        # Solo debe haber la fórmula legacy
        assert ws["B2"].value == "legacy_formula"
        # No debe haber v2_formula
        assert ws["B3"].value is None

    def test_engine_v2_resolves_input_references(self, tmp_path: Path):
        """Derived resuelve referencias a inputs correctamente."""
        import textwrap

        spec_file = tmp_path / "engine_v2_refs.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: engine_v2_refs
                title: "Test Engine V2 Refs"
                sheets:
                  - name: DATOS
                    type: input
                    inputs:
                      - id: precio
                        label: "Precio"
                        type: currency
                        default: 100
                      - id: costo
                        label: "Costo"
                        type: currency
                        default: 60
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    derived:
                      - id: margen_pesos
                        formula: margen_bruto_pesos
                        inputs:
                          precio_venta: precio
                          costo_unitario: costo
                  - name: DASHBOARD
                    type: dashboard
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "engine_v2_refs.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["MOTOR"]

        # Verificar que la fórmula tiene las referencias correctas
        formula = str(ws["C2"].value)
        # precio está en DATOS!C5, costo está en DATOS!C6
        assert "DATOS!C5" in formula
        assert "DATOS!C6" in formula

    def test_engine_v2_resolves_derived_references(self, tmp_path: Path):
        """Derived puede referenciar otros derived."""
        import textwrap

        spec_file = tmp_path / "engine_v2_chain.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: engine_v2_chain
                title: "Test Engine V2 Chain"
                sheets:
                  - name: DATOS
                    type: input
                    inputs:
                      - id: costo
                        label: "Costo"
                        type: currency
                        default: 100
                      - id: margen
                        label: "Margen"
                        type: percentage
                        default: 0.40
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    derived:
                      - id: precio
                        formula: precio_venta_con_margen
                        inputs:
                          costo_unitario: costo
                          margen_objetivo: margen
                      - id: margen_bruto_pct
                        formula: margen_bruto
                        inputs:
                          precio_venta: precio
                          costo_unitario: costo
                  - name: DASHBOARD
                    type: dashboard
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "engine_v2_chain.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["MOTOR"]

        # Segunda fórmula debe referenciar MOTOR!C2 (precio)
        formula = str(ws["C3"].value)
        assert "MOTOR!C2" in formula

    def test_engine_v2_invalid_reference_fails(self, tmp_path: Path):
        """Referencia a input inexistente debe fallar."""
        import textwrap

        spec_file = tmp_path / "engine_v2_invalid.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: engine_v2_invalid
                title: "Test Engine V2 Invalid"
                sheets:
                  - name: DATOS
                    type: input
                    inputs:
                      - id: costo
                        label: "Costo"
                        type: currency
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    derived:
                      - id: margen
                        formula: margen_bruto
                        inputs:
                          precio_venta: no_existe
                          costo_unitario: costo
                  - name: DASHBOARD
                    type: dashboard
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "engine_v2_invalid.xlsx"
        result = build_workbook(spec, out)

        # Debe fallar con error claro
        assert not result.success
        assert "no_existe" in result.error or "no existe" in result.error

    def test_legacy_engine_still_works(self, tmp_path: Path):
        """Specs legacy con formulas siguen funcionando."""
        from exceland_factory.factory import build_product

        out = tmp_path / "caja_diaria.xlsx"
        result = build_product("caja_diaria", out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["MOTOR"]

        # Verificar que las fórmulas legacy están
        assert ws["B1"].value == "ID"
        assert ws["C1"].value == "FÓRMULA"
