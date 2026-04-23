"""Tests de DASHBOARD v2: validar renderizado desde outputs declarativos."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from exceland_factory.workbook_builder import build_workbook
from exceland_factory.validators import load_spec


class TestDashboardV2Rendering:
    def test_dashboard_v2_renders_from_outputs_field(self, tmp_path: Path):
        """Si no hay kpis legacy pero hay outputs, renderiza desde outputs."""
        import textwrap

        spec_file = tmp_path / "dashboard_v2.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: dashboard_v2_test
                title: "Test Dashboard V2"
                sheets:
                  - name: DATOS
                    type: input
                    inputs:
                      - id: costo
                        label: "Costo"
                        type: currency
                        default: 100
                      - id: margen
                        label: "Margen"
                        type: percentage
                        default: 0.40
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    derived:
                      - id: precio_venta
                        formula: precio_venta_con_margen
                        inputs:
                          costo_unitario: costo
                          margen_objetivo: margen
                      - id: margen_bruto
                        formula: margen_bruto
                        inputs:
                          precio_venta: precio_venta
                          costo_unitario: costo
                  - name: DASHBOARD
                    type: dashboard
                    protected: true
                    outputs:
                      - id: kpi_precio
                        label: "Precio de venta"
                        source: precio_venta
                        style: kpi_positive
                      - id: kpi_margen
                        label: "Margen bruto %"
                        source: margen_bruto
                        style: kpi_neutral
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "dashboard_v2.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        # Verificar que los KPIs se renderizaron
        assert ws["B5"].value == "Precio de venta"
        assert ws["B7"].value == "Margen bruto %"

        # Verificar que las fórmulas apuntan a MOTOR
        formula_precio = str(ws["C5"].value)
        formula_margen = str(ws["C7"].value)
        assert "MOTOR!C2" in formula_precio  # precio_venta está en MOTOR!C2
        assert "MOTOR!C3" in formula_margen  # margen_bruto está en MOTOR!C3

    def test_dashboard_legacy_takes_precedence_over_v2(self, tmp_path: Path):
        """Si hay kpis (legacy) y outputs (v2), kpis tiene precedencia."""
        import textwrap

        spec_file = tmp_path / "mixed_dashboard.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: mixed_dashboard_test
                title: "Test Mixed Dashboard"
                sheets:
                  - name: DATOS
                    type: input
                    fields:
                      - id: x
                        label: "X"
                        row: 5
                        col: 2
                        input_type: text
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    formulas:
                      - id: legacy_formula
                        formula_ref: margen_bruto
                        bindings:
                          precio_venta: DATOS!C5
                          costo_unitario: DATOS!C6
                  - name: DASHBOARD
                    type: dashboard
                    protected: true
                    kpis:
                      - label: "Legacy KPI"
                        ref: "MOTOR!C2"
                        style: kpi_positive
                    outputs:
                      - id: v2_kpi
                        label: "V2 KPI"
                        source: legacy_formula
                        style: kpi_neutral
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "mixed_dashboard.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        # Solo debe haber el KPI legacy
        assert ws["B5"].value == "Legacy KPI"
        # No debe haber "V2 KPI" en ninguna parte del dashboard
        all_values = [ws.cell(row=r, column=2).value for r in range(1, 20)]
        assert "V2 KPI" not in [v for v in all_values if v is not None]

    def test_dashboard_v2_resolves_source_against_derived(self, tmp_path: Path):
        """Outputs resuelve source contra derived correctamente."""
        import textwrap

        spec_file = tmp_path / "dashboard_v2_source.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: dashboard_v2_source
                title: "Test Dashboard V2 Source"
                sheets:
                  - name: DATOS
                    type: input
                    inputs:
                      - id: precio
                        label: "Precio"
                        type: currency
                        default: 100
                      - id: costo
                        label: "Costo"
                        type: currency
                        default: 60
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    derived:
                      - id: margen_pesos
                        formula: margen_bruto_pesos
                        inputs:
                          precio_venta: precio
                          costo_unitario: costo
                  - name: DASHBOARD
                    type: dashboard
                    protected: true
                    outputs:
                      - id: kpi_margen
                        label: "Margen en pesos"
                        source: margen_pesos
                        style: kpi_positive
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "dashboard_v2_source.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        # Verificar que la fórmula apunta a MOTOR!C2
        formula = str(ws["C5"].value)
        assert "MOTOR!C2" in formula

    def test_dashboard_v2_invalid_source_fails(self, tmp_path: Path):
        """Source que no existe en derived debe fallar."""
        import textwrap

        spec_file = tmp_path / "dashboard_v2_invalid.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: dashboard_v2_invalid
                title: "Test Dashboard V2 Invalid"
                sheets:
                  - name: DATOS
                    type: input
                    inputs:
                      - id: costo
                        label: "Costo"
                        type: currency
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    derived:
                      - id: precio
                        formula: precio_venta_con_margen
                        inputs:
                          costo_unitario: costo
                          margen_objetivo: costo
                  - name: DASHBOARD
                    type: dashboard
                    protected: true
                    outputs:
                      - id: kpi_invalid
                        label: "KPI Invalid"
                        source: no_existe
                        style: kpi_positive
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "dashboard_v2_invalid.xlsx"
        result = build_workbook(spec, out)

        # Debe fallar con error claro
        assert not result.success
        assert "no_existe" in result.error or "no existe" in result.error

    def test_dashboard_legacy_still_works(self, tmp_path: Path):
        """Specs legacy con kpis.ref siguen funcionando."""
        from exceland_factory.factory import build_product

        out = tmp_path / "caja_diaria.xlsx"
        result = build_product("caja_diaria", out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        # Verificar que los KPIs legacy están
        assert ws["B5"].value == "Ingresos totales"
        assert ws["B7"].value == "Egresos totales"

    def test_dashboard_fallback_when_no_kpis_or_outputs(self, tmp_path: Path):
        """Si no hay kpis ni outputs, usa fallback legacy."""
        import textwrap

        spec_file = tmp_path / "dashboard_fallback.yaml"
        spec_file.write_text(
            textwrap.dedent("""\
                slug: dashboard_fallback
                title: "Test Dashboard Fallback"
                sheets:
                  - name: DATOS
                    type: input
                    fields:
                      - id: x
                        label: "X"
                        row: 5
                        col: 2
                        input_type: text
                  - name: MOTOR
                    type: engine
                    protected: true
                    hidden: true
                    formulas: []
                  - name: DASHBOARD
                    type: dashboard
                    protected: true
                  - name: GUIA
                    type: guide
            """),
            encoding="utf-8",
        )
        spec = load_spec(spec_file)
        out = tmp_path / "dashboard_fallback.xlsx"
        result = build_workbook(spec, out)

        assert result.success, f"Build falló: {result.error}"

        wb = load_workbook(str(out))
        ws = wb["DASHBOARD"]

        # Verificar que el fallback legacy se aplicó
        assert ws["B5"].value == "Ingresos totales"
        assert ws["B7"].value == "Egresos totales"
        assert ws["B9"].value == "Resultado neto"
        assert ws["B11"].value == "Punto de equilibrio (uds)"
