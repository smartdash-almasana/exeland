"""Tests de build-all: todos los productos registrados."""
from __future__ import annotations

from pathlib import Path

import pytest

from exceland_factory.factory import build_all_products, build_product
from exceland_factory.registry import list_products, load_product_registry


class TestBuildAll:
    def test_build_all_returns_results(self, tmp_path: Path):
        results = build_all_products(tmp_path)
        assert isinstance(results, list)
        assert len(results) > 0

    def test_build_all_generates_four_products(self, tmp_path: Path):
        results = build_all_products(tmp_path)
        assert len(results) == 4

    def test_all_builds_succeed(self, tmp_path: Path):
        results = build_all_products(tmp_path)
        failures = [r for r in results if not r.success]
        assert failures == [], (
            f"Builds fallidos: {[(r.slug, r.error) for r in failures]}"
        )

    def test_all_files_exist(self, tmp_path: Path):
        results = build_all_products(tmp_path)
        for r in results:
            assert Path(r.output_path).exists(), (
                f"Archivo no generado: {r.output_path}"
            )

    def test_all_files_are_valid_xlsx(self, tmp_path: Path):
        """Verifica que cada archivo sea un xlsx abrible con openpyxl."""
        from openpyxl import load_workbook
        results = build_all_products(tmp_path)
        for r in results:
            assert r.success
            wb = load_workbook(r.output_path)
            assert len(wb.sheetnames) >= 4, (
                f"{r.slug} debería tener al menos 4 hojas"
            )

    def test_slugs_match_registry(self, tmp_path: Path):
        registered = set(list_products())
        results = build_all_products(tmp_path)
        built = {r.slug for r in results}
        assert built == registered

    def test_build_all_without_output_dir(self):
        """Sin output_dir usa dist/ por defecto."""
        results = build_all_products(None)
        for r in results:
            assert r.success, f"Falló: {r.slug} — {r.error}"
            assert Path(r.output_path).exists()

    def test_each_product_individually(self, tmp_path: Path):
        """Verifica cada producto por separado."""
        slugs = ["caja_diaria", "precio_margen", "stock_control", "punto_equilibrio"]
        for slug in slugs:
            out = tmp_path / f"{slug}.xlsx"
            result = build_product(slug, out)
            assert result.success, f"Falló '{slug}': {result.error}"
            assert out.exists(), f"Archivo no encontrado: {out}"

    def test_precio_margen_has_correct_sheets(self, tmp_path: Path):
        from openpyxl import load_workbook
        out = tmp_path / "precio_margen.xlsx"
        build_product("precio_margen", out)
        wb = load_workbook(str(out))
        assert "DATOS" in wb.sheetnames
        assert "DASHBOARD" in wb.sheetnames

    def test_stock_control_has_correct_sheets(self, tmp_path: Path):
        from openpyxl import load_workbook
        out = tmp_path / "stock_control.xlsx"
        build_product("stock_control", out)
        wb = load_workbook(str(out))
        assert "STOCK" in wb.sheetnames

    def test_punto_equilibrio_has_correct_sheets(self, tmp_path: Path):
        from openpyxl import load_workbook
        out = tmp_path / "punto_equilibrio.xlsx"
        build_product("punto_equilibrio", out)
        wb = load_workbook(str(out))
        assert "DATOS" in wb.sheetnames
        assert "BIENVENIDA" in wb.sheetnames

    def test_registry_has_four_entries(self):
        registry = load_product_registry()
        assert len(registry) == 4

    def test_registry_entries_have_spec_paths(self):
        registry = load_product_registry()
        for slug, entry in registry.items():
            assert entry.spec_path, f"spec_path vacío en '{slug}'"
            # precio puede ser 0 para productos free (ej: punto_equilibrio)
            assert entry.price_ars >= 0, f"price_ars no puede ser negativo en '{slug}'"
