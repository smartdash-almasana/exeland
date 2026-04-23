"""Postprocess: aplica validaciones de celda con openpyxl tras generar el xlsx."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from exceland_factory.models import InputType, ProductSpec, SheetType
from exceland_factory.registry import get_validation


def apply_validations(spec: ProductSpec, xlsx_path: Path) -> None:
    """
    Abre el xlsx generado con openpyxl y aplica DataValidation
    a cada campo que tenga un validation definido en su FieldSpec o InputSpecV2.

    Se ejecuta después de build_workbook() porque xlsxwriter no soporta
    agregar validaciones a celdas con datos ya escritos de forma explícita.
    """
    wb = load_workbook(str(xlsx_path))

    for sheet_spec in spec.sheets:
        if sheet_spec.type != SheetType.input:
            continue
        if sheet_spec.name not in wb.sheetnames:
            continue

        ws = wb[sheet_spec.name]

        # Procesar fields (legacy)
        for fspec in sheet_spec.fields:
            if not fspec.validation:
                continue

            val_def = get_validation(fspec.validation)
            if val_def is None:
                continue

            # Construir la celda de destino (col es 1-indexed en spec, A=1)
            col_letter = _col_letter(fspec.col + 1)  # col en spec es 0-indexed desde A=0
            cell_ref = f"{col_letter}{fspec.row}"

            dv = DataValidation(
                type=val_def.type,
                operator=val_def.operator,
                formula1=val_def.formula1,
                formula2=val_def.formula2,
                showErrorMessage=val_def.show_error,
                errorTitle=val_def.error_title,
                error=val_def.error_message,
            )
            dv.sqref = cell_ref
            ws.add_data_validation(dv)

        # Procesar inputs (v2) - layout automático: columna C, filas secuenciales desde 5
        start_row = 5
        for i, inp in enumerate(sheet_spec.inputs):
            if not inp.validation:
                continue

            val_def = get_validation(inp.validation)
            if val_def is None:
                continue

            # Layout automático: columna C (índice 2), fila = start_row + i
            cell_ref = f"C{start_row + i}"

            dv = DataValidation(
                type=val_def.type,
                operator=val_def.operator,
                formula1=val_def.formula1,
                formula2=val_def.formula2,
                showErrorMessage=val_def.show_error,
                errorTitle=val_def.error_title,
                error=val_def.error_message,
            )
            dv.sqref = cell_ref
            ws.add_data_validation(dv)

    wb.save(str(xlsx_path))


def _col_letter(n: int) -> str:
    """Convierte número de columna (1-indexed) a letra Excel. Ej: 1→A, 3→C."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result
